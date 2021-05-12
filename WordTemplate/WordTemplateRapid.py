import os
from pathlib import Path
import shutil
from docxtpl import DocxTemplate
import configparser
from docxcompose.composer import Composer
from docx import Document as Document_compose
import openpyxl, openpyxl.utils


def wordTemplateRapid(group, pattern):
    i = 0
    single_diploma = []

    lists_path = Path('lists')
    pattern_path = Path('patterns')
    diploma_path = Path('diplomas')
    spiski = []

    def spisok():
        for file in os.listdir(lists_path):
            spiski.append(file)

    def otchestvo():
        list_line = list(line)
        fio_len = len(list_line)
        ending = line[fio_len - 2:fio_len]
        if ending == 'ич':
            return 'прошел'
        elif ending == 'на':
            return 'прошла'
        else:
            print(prog, ': TROUBLES WITH ' + line)

    class Config:
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='UTF-8')
        date1 = config.get('Config', 'date1')
        date2 = config.get('Config', 'date2')
        director = config.get('Config', 'director')
        year = config.get('Config', 'year')

    spisok()
    shutil.rmtree("diplomas", ignore_errors=True)
    os.mkdir("diplomas")

    wb = openpyxl.load_workbook(lists_path / group)
    sheet = wb.active
    rows = sheet.max_row

    prog = sheet.cell(row=1, column=1).value
    name_len = len(prog)

    for row_num in range(2, rows + 1):
        if pattern == 'base' or 'project':
            if name_len < 85:
                patternName = pattern + '_1.docx'
                doc = DocxTemplate(pattern_path / patternName)
            elif (name_len > 84) and (name_len < 127):
                patternName = pattern + '_2.docx'
                doc = DocxTemplate(pattern_path / patternName)
            elif (name_len > 126) and (name_len < 161):
                patternName = pattern + '_3.docx'
                doc = DocxTemplate(pattern_path / patternName)
            elif name_len > 161:
                patternName = pattern + '_4.docx'
                doc = DocxTemplate(pattern_path / patternName)
        else:
            doc = DocxTemplate(pattern_path / pattern + '.docx')

        line = sheet.cell(row=row_num, column=1).value + ' ' + \
               sheet.cell(row=row_num, column=2).value + ' ' + \
               sheet.cell(row=row_num, column=3).value

        context = {'fio': line,
                   'date1': Config.date1,
                   'date2': Config.date2,
                   'kvant': str(prog),
                   'director': Config.director,
                   'year': Config.year,
                   'completed': otchestvo()}
        doc.render(context)
        doc.save("diploma_" + str(i) + ".docx")
        shutil.move("diploma_" + str(i) + ".docx", "diplomas")
        diploma_path = Path('diplomas/diploma_' + str(i) + '.docx')
        single_diploma.append(diploma_path)
        i += 1

    del single_diploma[0]
