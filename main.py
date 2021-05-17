import sys  # sys нужен для передачи argv в QApplication
import os  # Отсюда нам понадобятся методы для отображения содержимого директорий

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QAction

# import gui  # Это наш конвертированный файл дизайна
# import gui_v2  # Это наш конвертированный файл дизайна
# import gui_v3  # Это наш конвертированный файл дизайна
import gui4  # Это наш конвертированный файл дизайна

import shutil
from pathlib import Path
import comtypes.client
import win32com.client
import openpyxl
import openpyxl.utils
from docxtpl import DocxTemplate

templates = os.listdir('learn_templates')
user_list = os.listdir('user_lists')
wdFormatPDF = 17


def view_diplomas():
    os.system('explorer.exe "diplomas"')


def doc2pdf():
    """Convert a Word .docx to PDF"""

    print(os.listdir("diplomas"))
    docx_list = os.listdir("diplomas")
    os.chdir("diplomas")
    print(os.getcwd())
    i = 0
    for document in docx_list:
        word = comtypes.client.CreateObject('Word.Application')
        print(document)
        doc = word.Documents.Open(os.path.abspath(document))
        doc.SaveAs(os.getcwd() + '\\' + document + str(i) + '.pdf', FileFormat=wdFormatPDF)
        doc.Close()
        i += 1
        word.Quit()


def list_doc2pdf():  # "D:\\prj\\python_prj\\crn\\WordProgram\\diplomas"
    for root, dirs, files in os.walk(os.getcwd() + "\\diplomas"):
        for f in files:
            if f.endswith(".doc") or f.endswith(".odt") or f.endswith(".rtf"):
                try:
                    print(f)
                    in_file = os.path.join(root, f)
                    word = win32com.client.Dispatch('Word.Application')
                    word.Visible = False
                    doc = word.Documents.Open(in_file)
                    doc.SaveAs(os.path.join(root, f[:-4]), FileFormat=wdFormatPDF)
                    doc.Close()
                    word.Quit()
                    word.Visible = True
                    print('done')
                    # os.remove(os.path.join(root,f))
                    pass
                except:
                    pass
                    # print('could not open')
                    # os.remove(os.path.join(root,f))
            elif f.endswith(".docx") or f.endswith(".dotm") or f.endswith(".docm"):
                try:
                    print(f)
                    in_file = os.path.join(root, f)
                    word = win32com.client.Dispatch('Word.Application')
                    word.Visible = False
                    doc = word.Documents.Open(in_file)
                    doc.SaveAs(os.path.join(root, f[:-5]), FileFormat=wdFormatPDF)
                    doc.Close()
                    word.Quit()
                    word.Visible = True
                    print('done')
                    # os.remove(os.path.join(root,f))
                    pass
                except:
                    pass
                    # print('could not open')
                    # os.remove(os.path.join(root,f))
            else:
                pass


class ExampleApp(QtWidgets.QMainWindow, gui4.Ui_MainWindow):
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        self.pushButton_2.clicked.connect(self.start_generation)  # Выполнить функцию start_generation
        self.pushButton.clicked.connect(self.browse_folder)  # Выполнить функцию browse_folder
        self.pushButton_3.clicked.connect(self.view_diplomas)  # связь кнопки открыть папку с результатом (дипломами)
        self.comboBox.addItems(templates)  # заполнение comboBox списком файлов шаблонов
        self.comboBox_2.addItems(user_list)  # заполнение comboBox списком файлов групп
        self.dateEdit.setCalendarPopup(True)  # настройка даты опция выпадающего календаря
        self.dateEdit.setDateTime(QtCore.QDateTime.currentDateTime())  # устанавливаем текущую дату
        self.dateEdit_2.setDateTime(QtCore.QDateTime.currentDateTime())  # устанавливаем текущую дату
        self.dateEdit_2.setCalendarPopup(True)  # настройка даты опция выпадающего календаря
        self.progressBar.setValue(0)  # настройка первоначального значения progressBar
        self.menu.setEnabled(False)  # неактивный пункт меню "Справка"
        self.action = QAction("Выход")  # создание нового пункта меню
        self.menu_2.addAction(self.action)  # добавление в меню пункт "Выход"
        self.action.triggered.connect(self.close)  # кнопка закрытия приложения
        self.checkBox.setChecked(True)  # по умолчанию конвертация в pdf включена

    def view_diplomas(self):
        # lists_path = Path('diplomas')
        # print(lists_path)
        # list_doc2pdf()
        doc2pdf()
        # os.system('explorer.exe "diplomas"')

    def browse_folder(self):
        file_name = QtWidgets.QFileDialog.getOpenFileName(self, "Выбор шаблона", None, "word (*.doc *.docx)")[0]
        self.label_5.setText(file_name)

    def start_generation(self):
        template = str(self.comboBox.currentText())
        group_list = str(self.comboBox_2.currentText())
        # --- radio button --- #
        if self.radioButton.isChecked():
            print('Базовая группа')
        elif self.radioButton_2.isChecked():
            print('Проектная группа')
        elif self.radioButton_3.isChecked():
            print('Свой шаблон')
        # --- checkBox --- #
        if self.checkBox_2.isChecked():
            print('Сохранить в один файл')
        if self.checkBox_3.isChecked():
            print('Дата начала и окончания')

        self.label_5.setText("Готовим Docx...")
        lists_path = Path('user_lists')
        pattern_path = Path('learn_templates')
        shutil.rmtree("diplomas", ignore_errors=True)
        os.mkdir("diplomas")
        i = 0  # счетчик для именования файлов
        loading = 0  # значение для индикации загрузки
        wb = openpyxl.load_workbook(lists_path / group_list)
        sheet = wb.active
        rows = sheet.max_row
        step = 100 / rows  # вычисление шага преодразования одного документа для индикации
        # (первая строка из документа Excel)
        pattern_name = template  # название шаблона
        date1 = self.dateEdit.date().toString('dd.MM.yyyy')     # дата начала
        date2 = self.dateEdit_2.date().toString('dd.MM.yyyy')   # и окончания обучения
        duration = self.lineEdit.text()     # продолжительнсть учебной программы
        for row_num in range(2, rows + 1):
            context = {'kvant': str(sheet.cell(row=1, column=1).value), 'date1': date1, 'date2': date2, 'duration': duration}
            fio = str(sheet.cell(row=row_num, column=1).value) + ' ' + str(sheet.cell(row=row_num, column=2).value) + ' ' + str(sheet.cell(row=row_num, column=3).value)
            context.setdefault('fio', fio)
            loading += step
            doc = DocxTemplate(pattern_path / pattern_name)
            doc.render(context)
            name_document = fio + '_' + str(i) + ".docx"
            doc.save(name_document)
            shutil.move(name_document, "diplomas")
            i += 1
            self.progressBar.setValue(int(loading))
        self.progressBar.setValue(100)

        if self.checkBox.isChecked():  # если выбран пункт сохранить в формате pdf,
            # то все документы docx из папки diplomas переконвертируются в pdf
            self.label_5.setText("Готовим PDF...")
            # convert("diplomas/")
            # self.label_5.setText("Свежие PDF готовы!)")


def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
