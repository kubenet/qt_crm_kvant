import os
import shutil
from pathlib import Path

import openpyxl
import openpyxl.utils
from docxtpl import DocxTemplate

loading = 0


def generate_documents(template, group_list, date1, date2):
    lists_path = Path('user_lists')
    pattern_path = Path('learn_templates')
    shutil.rmtree("diplomas", ignore_errors=True)
    os.mkdir("diplomas")
    print(template, group_list, date1, date2)
    i = 0
    context = {}
    wb = openpyxl.load_workbook(lists_path / group_list)
    sheet = wb.active
    rows = sheet.max_row
    step = rows/100
    learn_program = sheet.cell(row=1, column=1).value
    pattern_name = template
    doc = DocxTemplate(pattern_path / pattern_name)
    for row_num in range(2, rows + 1):
        line = sheet.cell(row=row_num, column=1).value + ' ' + \
               sheet.cell(row=row_num, column=2).value + ' ' + \
               sheet.cell(row=row_num, column=3).value
        global loading
        loading += step
        print(loading)
        context['fio'] = line
        context['date1'] = date1
        context['date2'] = date2
        context['kvant'] = str(learn_program)
        doc.render(context)
        name_document = str(i) + "_" + str(sheet.cell(row=row_num, column=1).value) + ".docx"
        doc.save(name_document)
        shutil.move(name_document, "diplomas")
        i += 1



